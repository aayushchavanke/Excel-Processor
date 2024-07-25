import tkinter as tk
from tkinter import ttk
from tkinter import messagebox, filedialog
import pandas as pd
from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import PatternFill, Font
import os


class ExcelProcessorApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Excel Processor")

        # File path
        self.file_path = tk.StringVar()

        # Styling
        self.root.configure(bg='lightblue')
        self.root.geometry('800x600')

        # Create GUI elements
        self.create_widgets()

    def create_widgets(self):
        # Main frame
        main_frame = tk.Frame(self.root, bg='#2e2e2e', padx=20, pady=20)
        main_frame.pack(fill=tk.BOTH, expand=True)

        # Title label
        title_label = tk.Label(main_frame, text="Excel Processor", font=('Arial', 18, 'bold'), bg='#2e2e2e', fg='#F0EAD6')
        title_label.grid(row=0, column=0, columnspan=3, pady=10)


        # File selection widgets
        tk.Label(main_frame, text="Excel File:", bg='#2e2e2e', fg='#CECECE', font=('Arial', 12, 'bold'),background='#2e2e2e',foreground='#FFD700').grid(row=1, column=0, padx=10, pady=10)
        tk.Entry(main_frame, textvariable=self.file_path, width=50, font=('Arial', 12),bg='#CECECE').grid(row=1, column=1, padx=10, pady=10)
        tk.Button(main_frame, text="Browse", command=self.browse_file, font=('Arial', 12, 'bold'), bg='#FF8C00', fg='#FFFFFF').grid(row=1, column=2, padx=10, pady=10)

        # Process button
        process_button = tk.Button(main_frame, text="Process", command=self.process, font=('Arial', 14, 'bold'), bg='#1E90FF', fg='#FFFFFF')
        process_button.grid(row=2, column=1, columnspan=1, padx=10, pady=20)
        
        # Provide Inputs button
        inputs_button = tk.Button(main_frame, text="Provide Inputs", command=self.provide_inputs, font=('Arial', 12, 'bold'), bg='#32CD32', fg='white')
        inputs_button.grid(row=2, column=2, columnspan=1, padx=10, pady=20)

        # Status text area
        self.status_text = tk.Text(main_frame, height=10, width=70, font=('Arial', 12), bg='#2F4F4F',fg='#FFFFFF')
        self.status_text.grid(row=3, column=0, columnspan=3, padx=10, pady=10)
        self.status_text.config(state=tk.DISABLED)  # Disable editing
        

        # Button for Final Sheet (initially hidden)
        self.final_sheet_button = tk.Button(main_frame, text="Open Final Sheet", command=self.open_final_sheet, font=('Arial', 12, 'bold'), bg='#FF4500', fg='#FFFFFF')
        self.final_sheet_button.grid(row=4, column=0, columnspan=3, padx=10, pady=10)
        self.final_sheet_button.grid_remove()  # Initially hidden
        
        style = ttk.Style()
        style.theme_use('clam')  # Use a modern theme
        style.configure('TButton', font=('Helvetica Neue', 12, 'bold'), padding=10, background='#4CAF50', foreground='#FFFFFF')
        style.configure('Accent.TButton', font=('Helvetica Neue', 14, 'bold'), padding=10, background='#2196F3', foreground='#FFFFFF')
        style.map('TButton', background=[('active', '#45a049')])
        style.map('Accent.TButton', background=[('active', '#1976D2')])



    def browse_file(self):
        file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx *.xlsm *.xltx *.xltm")])
        self.file_path.set(file_path)

    def log_status(self, message):
        self.status_text.config(state=tk.NORMAL)  # Enable editing
        self.status_text.insert(tk.END, message + '\n')
        self.status_text.see(tk.END)
        self.status_text.config(state=tk.DISABLED)  # Disable editing

    def create_table(self, ws, df, table_name):
        headers = list(df.columns)
        ws.append(headers)  # Append headers as the first row

        # Apply orange color to headers with bold and #2e2e2e font
        header_fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
        header_font = Font(color="000000", bold=True)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font

        # Write data rows
        for idx, row in df.iterrows():
            ws.append([row[col] for col in df.columns])
            
        # Apply Arial font with size 9 to all cells in the table
        data_font = Font(name="Arial", size=9)
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=len(df.columns)):
            for cell in row:
                cell.font = data_font


        # Add table and style
        table_range = f'A1:{chr(ord("A") + len(df.columns) - 1)}{len(df) + 1}'  # Adjust range for table
        table = Table(displayName=table_name, ref=table_range)

        # Define table style
        style = TableStyleInfo(
            name="TableStyle1",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=True
        )
        table.tableStyleInfo = style

        # Add table to worksheet
        ws.add_table(table)

    def create_model_counts(self, wb, file):
        try:
            df = pd.read_excel(file, sheet_name='BOM_Master')
            unique_df = df.drop_duplicates(subset=['Model', 'Main item trim'], keep='first').copy()
            unique_df['No. of Units'] = 0
            columns_to_keep = ['Family', 'Model', 'Main item trim', 'No. of Units']
            unique_df = unique_df[columns_to_keep]

            sheet_name = 'Model_Counts'
            if sheet_name in wb.sheetnames:
                del wb[sheet_name]

            ws = wb.create_sheet(title=sheet_name)
            self.create_table(ws, unique_df, sheet_name)
            wb.save(file)
            self.log_status(f"{sheet_name} created successfully.")
        except Exception as e:
            raise RuntimeError(f"Error creating {sheet_name}: {str(e)}")

    def create_stocks_analysis(self, wb, file):
        try:
            # Read the 'Item Trim' column from BOM_Master
            item_trim_df = pd.read_excel(file, sheet_name='BOM_Master', usecols=['Item Trim'])
            item_trim_values = item_trim_df['Item Trim'].unique()
            
            # Read data from Stock_input
            use_cols = ['Item Trim', 'Component_Desc', 'Stock_on_hand']
            df = pd.read_excel(file, sheet_name='Stock_input', usecols=use_cols)
            
            # Filter the dataframe to include only rows where 'Item Trim' matches those in BOM_Master
            df = df[df['Item Trim'].isin(item_trim_values)]
            
            # Add the 'Arrivals' column
            df['Arrivals'] = 0
            
            
            
            # Create the pivot table
            pivot_table = df.pivot_table(
                index=['Item Trim', 'Component_Desc'],
                values=['Stock_on_hand', 'Arrivals'],
                aggfunc='sum',
                fill_value=0
            )
            
            # Create or replace the 'Stocks_analysis' sheet
            sheet_name = 'Stocks_analysis'
            if sheet_name in wb.sheetnames:
                del wb[sheet_name]
                
            ws = wb.create_sheet(title=sheet_name)
            
            
            # Reset the index of the pivot table and write it to the new sheet
            pivot_table.reset_index(inplace=True)
            
            pivot_table = pivot_table[['Item Trim', 'Component_Desc', 'Stock_on_hand', 'Arrivals']]


            # Add headers
            headers = list(pivot_table.columns)
            ws.append(headers)  # Append headers as the first row

            # Apply orange color to headers with bold and #2e2e2e font
            header_fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
            header_font = Font(color="000000", bold=True)
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font

            for r_idx, row in pivot_table.iterrows():
                for c_idx, value in enumerate(row):
                    ws.cell(row=r_idx+2, column=c_idx+1, value=value)
            
            self.create_table(ws, pivot_table, sheet_name)

            # Save the workbook
            wb.save(file)
            self.log_status(f"{sheet_name} created successfully.")
        except Exception as e:
            raise RuntimeError(f"Error creating {sheet_name}: {str(e)}")

    def create_final_sheet(self, wb, file):
        try:
            use_col1 = ['Family', 'Model', 'Item Trim', 'Component_Desc', 'Qty_Per_Assly','Buyer Name','Supplier Name']
            use_col2 = ['Item Trim', 'Stock_on_hand', 'Arrivals']
            df_sheet1 = pd.read_excel(file, sheet_name='BOM_Master', usecols=use_col1)
            df_model_counts = pd.read_excel(file, sheet_name='Model_Counts')
            df_stocks_analysis = pd.read_excel(file, sheet_name='Stocks_analysis', usecols=use_col2)

            df_merged = pd.merge(df_sheet1, df_stocks_analysis, on='Item Trim', how='left')
            df_merged = pd.merge(df_merged, df_model_counts[['Model', 'No. of Units']], on='Model', how='left')

            df_merged['Total Requirement'] = df_merged['Qty_Per_Assly'] * df_merged['No. of Units']
            df_merged['Shortage/Excess'] = df_merged['Stock_on_hand'] + df_merged['Arrivals'] - df_merged['Total Requirement']
            df_merged['Remarks'] = ''
            
            if 'Model_Counts' in wb.sheetnames :
             model_counts_ws = wb['Model_Counts']
             model_counts_ws.sheet_properties.tabColor = "FFFF00"  # Yellow

            if 'Stocks_analysis' in wb.sheetnames:
             stocks_analysis_ws = wb['Stocks_analysis']
             stocks_analysis_ws.sheet_properties.tabColor = "FFFF00"  

            if 'Final_Sheet' in wb.sheetnames:
                del wb['Final_Sheet']

            ws = wb.create_sheet(title='Final_Sheet')
            headers = ['Family', 'Model', 'Item Trim', 'Component_Desc','Buyer Name','Supplier Name', 'Qty_Per_Assly', 'No. of Units',
                       'Total Requirement','Stock_on_hand','Arrivals','Shortage/Excess', 'Remarks']
            
            for col_num, header in enumerate(headers, 1):
                ws.cell(row=1, column=col_num, value=header)
                cell = ws.cell(row=1, column=col_num)
                cell.fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
                cell.font = Font(color="000000", bold=True)

            for r_idx, row in df_merged.iterrows():
                ws.cell(row=r_idx + 2, column=1, value=row['Family'])
                ws.cell(row=r_idx + 2, column=2, value=row['Model'])
                ws.cell(row=r_idx + 2, column=3, value=row['Item Trim'])
                ws.cell(row=r_idx + 2, column=4, value=row['Component_Desc'])
                ws.cell(row=r_idx + 2, column=5, value=row['Buyer Name'])
                ws.cell(row=r_idx + 2, column=6, value=row['Supplier Name'])
                ws.cell(row=r_idx + 2, column=7, value=row['Qty_Per_Assly'])
                ws.cell(row=r_idx + 2, column=8, value=row['No. of Units'])
                ws.cell(row=r_idx + 2, column=9, value=row['Total Requirement'])
                ws.cell(row=r_idx + 2, column=10, value=row['Stock_on_hand'])
                ws.cell(row=r_idx + 2, column=11, value=row['Arrivals'])
                ws.cell(row=r_idx + 2, column=12, value=row['Shortage/Excess'])
                ws.cell(row=r_idx + 2, column=13, value=row['Remarks'])
             
            # Add conditional formatting
            red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

            for row in range(2, len(df_merged) + 2):
                cell = ws.cell(row=row, column=headers.index('Shortage/Excess') + 1)
                value = cell.value
                if value < 0:
                    cell.fill = red_fill
                elif value > 0:
                    cell.fill = green_fill
                else:
                    cell.fill = yellow_fill

                

            table_range = f'A1:M{ws.max_row}'
            table = Table(displayName="Final_Sheet", ref=table_range)
            
            # Apply Arial font with size 9 to all cells in the table
            data_font = Font(name="Arial", size=9)
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=len(df_merged.columns)):
                for cell in row:
                    cell.font = data_font

            style = TableStyleInfo(
                name="TableStyle1",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=True
            )
            table.tableStyleInfo = style
            ws.add_table(table)
            ws.sheet_properties.tabColor = "00FF00" 
            wb.save(file)
            self.log_status("Final_Sheet created successfully.")
            self.log_status("----------------------------------------------------------------------")
            
            
            # Show the button to open Final Sheet
            self.final_sheet_button.grid(row=4, column=0, columnspan=3, padx=10, pady=10)

        except Exception as e:
            raise RuntimeError(f"Error creating Final_Sheet: {str(e)}")


    def process(self):
        try:
            file = self.file_path.get()
            if not file:
                messagebox.showerror("Error", "Please select an Excel file.")
                return

            if not os.path.exists(file):
                messagebox.showerror("Error", "File not found.")
                return

            wb = load_workbook(file)

            if 'Model_Counts' not in wb.sheetnames:
                self.create_model_counts(wb, file)
                self.log_status('Model_Counts sheet not found. Generating Model_Counts sheet...')

            if 'Stocks_analysis' not in wb.sheetnames:
                self.create_stocks_analysis(wb, file)
                self.log_status('Stocks_analysis sheet not found. Generating Stocks_analysis sheet...')


            self.create_final_sheet(wb, file)
            self.log_status("----------------------------------------------------------------------")

        except Exception as e:
            messagebox.showerror("Error", str(e))
            self.log_status(f"Error: {str(e)}")

    def open_final_sheet(self):
        file = self.file_path.get()
        if file and os.path.exists(file):
            os.system(f'start EXCEL.EXE "{file}"')
        else:
            messagebox.showerror("Error", "File not found.")


    def provide_inputs(self):
        file = self.file_path.get()
        if not file:
            messagebox.showerror("Error", "Please select an Excel file.")
            return

        if not os.path.exists(file):
            messagebox.showerror("Error", "The selected file does not exist.")
            return

        try:
            wb = load_workbook(file)
            self.create_model_counts(wb, file)
            self.create_stocks_analysis(wb, file)
        except Exception as e:
            messagebox.showerror("Error",(str(e)))
            self.log_status(str(e))
            
        if file and os.path.exists(file):
             os.system(f'start EXCEL.EXE "{file}"')
        else:
             messagebox.showerror("Error", "File not found.")
            
            
       

        self.log_status("----------------------------------------------------------------------")

# Create the main window and run the app
if __name__ == "__main__":
    root = tk.Tk()
    app = ExcelProcessorApp(root)
    root.mainloop()
